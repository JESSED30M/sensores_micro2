
import serial
import threading
import tkinter as tk
from tkinter import messagebox
import pandas as pd
import time
import os
from openpyxl import load_workbook

# === Configuración del puerto ===
puerto = 'COM6'         # Cambiar según tu configuración
baudrate = 9600
arduino = None
lectura_activa = False
datos_guardados = []

# === Función para leer datos del Arduino ===
def leer_serial():
    global lectura_activa
    buffer = []
    while lectura_activa:
        try:
            if arduino.in_waiting:
                linea = arduino.readline().decode('utf-8', errors='ignore').strip()

                if "LM35" in linea:
                    buffer = [linea]  # reiniciar buffer con LM35

                    # Capturar las siguientes 3 líneas
                    while len(buffer) < 4:
                        if arduino.in_waiting:
                            nueva_linea = arduino.readline().decode('utf-8', errors='ignore').strip()
                            if nueva_linea:
                                buffer.append(nueva_linea)

                    # Mostrar en consola con separación clara
                    print("======== MEDICIÓN ========")
                    for l in buffer:
                        print(l)
                    print("==========================\n")

                    # Actualizar interfaz y guardar datos
                    actualizar_interfaz(buffer)

        except Exception as e:
            print("Error al leer datos del puerto:", e)

        time.sleep(0.1)

# === Función para actualizar etiquetas de la GUI ===
def actualizar_interfaz(lineas):
    datos = {"LM35": "---", "DHT": "---", "Humedad": "---", "Distancia": "---"}
    for l in lineas:
        if "LM35" in l:
            datos["LM35"] = l.split(":")[1].strip()
        elif "DHT11 Temp" in l:
            datos["DHT"] = l.split(":")[1].strip()
        elif "Humedad" in l:
            datos["Humedad"] = l.split(":")[1].strip()
        elif "Distancia" in l:
            datos["Distancia"] = l.split(":")[1].strip()

    # Actualizar etiquetas visuales
    label_lm35.config(text=f"LM35 (°C): {datos['LM35']}")
    label_dht.config(text=f"DHT11 Temp (°C): {datos['DHT']}")
    label_hum.config(text=f"Humedad (%): {datos['Humedad']}")
    label_dist.config(text=f"Distancia (cm): {datos['Distancia']}")

    # Guardar en lista para exportar (con tiempo actual)
    datos_guardados.append([
        time.strftime("%Y-%m-%d %H:%M:%S"),
        datos["LM35"],
        datos["DHT"],
        datos["Humedad"],
        datos["Distancia"]
    ])

# === Funciones de los botones ===
def iniciar():
    global arduino, lectura_activa
    try:
        if not arduino:
            arduino = serial.Serial(puerto, baudrate, timeout=1)
            time.sleep(2)  # espera a que el Arduino reinicie
        lectura_activa = True
        threading.Thread(target=leer_serial, daemon=True).start()
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir {puerto}.\n{e}")

def pausar():
    global lectura_activa
    lectura_activa = False

def guardar_excel():
    nombre_archivo = "datos_sensores.xlsx"
    df = pd.DataFrame(datos_guardados, columns=["FechaHora", "LM35 (°C)", "DHT11 (°C)", "Humedad (%)", "Distancia (cm)"])

    try:
        if os.path.exists(nombre_archivo):
            # Abrir archivo existente
            libro = load_workbook(nombre_archivo)

            # Crear nombre para nueva hoja con timestamp
            nueva_hoja = time.strftime("Medicion_%Y%m%d_%H%M%S")

            # Convertir DataFrame a lista para escribir
            datos_excel = [df.columns.values.tolist()] + df.values.tolist()

            # Crear nueva hoja
            hoja = libro.create_sheet(title=nueva_hoja)

            # Escribir datos en la hoja
            for fila_idx, fila_datos in enumerate(datos_excel, start=1):
                for col_idx, valor in enumerate(fila_datos, start=1):
                    hoja.cell(row=fila_idx, column=col_idx, value=valor)

            # Guardar libro con la nueva hoja
            libro.save(nombre_archivo)

        else:
            # Guardar nuevo archivo si no existe
            df.to_excel(nombre_archivo, index=False)

        messagebox.showinfo("Guardado", f"Datos guardados en '{nombre_archivo}'")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar en Excel.\n{e}")

# === Interfaz Gráfica ===
ventana = tk.Tk()
ventana.title("Monitor de Sensores Arduino")
ventana.configure(bg="#f0f4f8")
ventana.geometry("420x380")

estilo_label = {"font": ("Arial", 14, "bold"), "bg": "#f0f4f8"}
estilo_titulo = {"font": ("Arial", 16, "bold"), "bg": "#ddeefc", "fg": "#0a3d62", "padx": 10, "pady": 10}

tk.Label(ventana, text="Lectura de Sensores", **estilo_titulo).pack(pady=10, fill="x")

label_lm35 = tk.Label(ventana, text="LM35 (°C): ---", fg="#e74c3c", **estilo_label)
label_dht = tk.Label(ventana, text="DHT11 Temp (°C): ---", fg="#3498db", **estilo_label)
label_hum = tk.Label(ventana, text="Humedad (%): ---", fg="#16a085", **estilo_label)
label_dist = tk.Label(ventana, text="Distancia (cm): ---", fg="#9b59b6", **estilo_label)

label_lm35.pack(pady=5)
label_dht.pack(pady=5)
label_hum.pack(pady=5)
label_dist.pack(pady=5)

# Botones con estilo
tk.Button(ventana, text="Iniciar", command=iniciar, bg="#2ecc71", fg="white", font=("Arial", 12), width=15).pack(pady=5)
tk.Button(ventana, text="Pausar", command=pausar, bg="#f39c12", fg="white", font=("Arial", 12), width=15).pack(pady=5)
tk.Button(ventana, text="Guardar Excel", command=guardar_excel, bg="#3498db", fg="white", font=("Arial", 12), width=15).pack(pady=5)

ventana.mainloop()






